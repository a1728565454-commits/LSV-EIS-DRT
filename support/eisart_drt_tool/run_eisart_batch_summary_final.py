# -*- coding: utf-8 -*-
from __future__ import annotations

import builtins
import importlib
import importlib.util
import os
import re
import shutil
import sys
import tempfile
import traceback
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox

AREA_CM2 = 0.196
BASE_DIR = Path(__file__).resolve().parent
DEFAULT_EISART_CODE_DIR = BASE_DIR / 'code'
KEEP_FOLDERS = [
    "DRT",
    "ECM",
    "DRT_ECM",
    "EIS_ECM",
    "EIS_Residual",
    "EIS_Weights",
    "EISART_Plots",
]
SUPPORTED_EXTS = {".txt", ".csv", ".ism"}
HEADER_COLORS = [
    "FFFF00", "FFC000", "92D050", "00B0F0",
    "D9E1F2", "DDD9C4", "F4B084", "A9D18E",
]
NUM_PATTERN = re.compile(r"(?<![A-Za-z])(\d+(?:\.\d+)?)")


def show_info(msg: str, title: str = "提示"):
    root = tk.Tk()
    root.withdraw()
    messagebox.showinfo(title, msg)
    root.destroy()


def show_error(msg: str, title: str = "出错"):
    root = tk.Tk()
    root.withdraw()
    messagebox.showerror(title, msg)
    root.destroy()


def choose_eis_files() -> list[Path]:
    root = tk.Tk()
    root.withdraw()
    files = filedialog.askopenfilenames(
        title="选择要处理的 EIS 文件（可多选）",
        filetypes=[
            ("EIS files", "*.txt *.csv *.ism"),
            ("Text files", "*.txt"),
            ("CSV files", "*.csv"),
            ("ISM files", "*.ism"),
            ("All files", "*.*"),
        ],
    )
    root.destroy()
    return [Path(f) for f in files]


def choose_directory(title: str) -> Path | None:
    root = tk.Tk()
    root.withdraw()
    folder = filedialog.askdirectory(title=title)
    root.destroy()
    return Path(folder) if folder else None


def _looks_like_numeric_row(parts: list[str]) -> bool:
    nums = 0
    for x in parts:
        try:
            float(x)
            nums += 1
        except Exception:
            pass
    return nums >= 3


def is_probable_eis_file(path: Path) -> tuple[bool, str]:
    """
    自动识别是否像原始 EIS 输入文件，避免把 summary、DRT、拟合结果等 txt/csv 当成原始 EIS 再处理。
    返回 (是否接受, 原因)
    """
    suffix = path.suffix.lower()
    stem_lower = path.stem.lower()
    name_lower = path.name.lower()

    if suffix == ".ism":
        return True, "ISM 原始阻抗文件"

    blocked_name_keywords = [
        "summary", "drt_ecm", "eis_ecm", "eis_residual", "eis_weights",
        "eisart_plots", "gamma", "residual", "weights", "fit", "plot"
    ]
    if any(k in name_lower for k in blocked_name_keywords):
        return False, "文件名看起来不是原始 EIS，而像导出结果/汇总文件"

    try:
        raw = path.read_text(encoding="utf-8", errors="ignore")
    except Exception:
        try:
            raw = path.read_text(encoding="gbk", errors="ignore")
        except Exception:
            return False, "无法读取文本内容"

    lines = [ln.strip() for ln in raw.splitlines() if ln.strip()]
    if not lines:
        return False, "空文件"

    header_blob = " ".join(lines[:10]).lower()
    if any(k in header_blob for k in ["gamma", "tau", "drt-fit-summary", "weighted eis fit rms error"]):
        return False, "内容看起来像 DRT/拟合导出文件，不是原始 EIS"

    if ("freq" in header_blob or "frequency" in header_blob) and ("imag" in header_blob or "zimag" in header_blob or "z''" in header_blob):
        return True, "表头包含频率和阻抗列"

    numeric_rows = 0
    first_col = []
    for ln in lines[:120]:
        parts = [x for x in re.split(r"[\s,\t;,_]+", ln) if x]
        if len(parts) < 3:
            continue
        vals = []
        for x in parts[:4]:
            try:
                vals.append(float(x))
            except Exception:
                pass
        if len(vals) >= 3:
            numeric_rows += 1
            first_col.append(vals[0])

    if numeric_rows < 8:
        return False, "不像原始 EIS 数据表（有效数值行过少或列数不足）"

    positive_freq = [x for x in first_col if x > 0]
    if len(positive_freq) < max(5, len(first_col) // 2):
        return False, "第一列不像频率列"

    monotonic_pairs = sum(
        1 for i in range(len(positive_freq) - 1) if positive_freq[i] != positive_freq[i + 1]
    )
    if monotonic_pairs == 0:
        return False, "频率列没有变化"

    return True, "数值结构看起来像原始 EIS"


def filter_selected_input_files(files: list[Path]) -> tuple[list[Path], list[str]]:
    accepted: list[Path] = []
    skipped: list[str] = []
    for f in files:
        if f.suffix.lower() not in SUPPORTED_EXTS:
            skipped.append(f"{f.name} -> 扩展名不支持")
            continue
        ok, reason = is_probable_eis_file(f)
        if ok:
            accepted.append(f)
        else:
            skipped.append(f"{f.name} -> 已跳过：{reason}")
    return accepted, skipped


def find_eisart_code_dir() -> Path:
    candidates = [
        DEFAULT_EISART_CODE_DIR,
        Path.cwd() / 'code',
    ]
    for c in candidates:
        if c.exists() and (c / 'EISART_kernel.py').exists():
            return c
    raise FileNotFoundError(
        '没找到 EISART 的 code 文件夹。请确认 run_eisart_batch_summary_final.py 与 code 文件夹放在同一目录。'
    )


def validate_eisart_code_dir(code_dir: Path):
    required = [
        "EISART_kernel.py", "util_io.py", "util_plot.py", "util_tikhonov.py", "util_ecm.py"
    ]
    missing = [name for name in required if not (code_dir / name).exists()]
    if missing:
        raise FileNotFoundError(
            "固定的 EISART code 文件夹无效，缺少这些文件：\n" + "\n".join(missing)
        )


def find_settings_dir(code_dir: Path) -> Path:
    for p in (code_dir, code_dir.parent):
        if (p / "settings_eisart.txt").exists():
            return p
    raise FileNotFoundError("没找到 settings_eisart.txt，请确认 EISART 安装完整。")


def safe_rmtree(path: Path):
    if path.exists():
        shutil.rmtree(path, ignore_errors=True)


def safe_copytree(src: Path, dst: Path):
    if dst.exists():
        shutil.rmtree(dst, ignore_errors=True)
    shutil.copytree(src, dst)


def patch_eisart_source_compatibility(code_dir: Path):
    """自动修正常见旧代码兼容性问题。"""
    util_ecm = code_dir / "util_ecm.py"
    if util_ecm.exists():
        text = util_ecm.read_text(encoding="utf-8", errors="ignore")
        new_text = text.replace("dtype=np.int)", "dtype=int)")
        new_text = new_text.replace("dtype = np.int)", "dtype=int)")
        if new_text != text:
            util_ecm.write_text(new_text, encoding="utf-8")

    util_plot = code_dir / "util_plot.py"
    if util_plot.exists():
        text = util_plot.read_text(encoding="utf-8", errors="ignore")
        new_text = text.replace("figize=", "figsize=")
        new_text = new_text.replace("legendHandles", "legend_handles")
        if new_text != text:
            util_plot.write_text(new_text, encoding="utf-8")


def load_module_from_file(module_name: str, py_file: Path):
    spec = importlib.util.spec_from_file_location(module_name, str(py_file))
    if spec is None or spec.loader is None:
        raise ImportError(f"无法加载模块: {py_file}")
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return module


def monkeypatch_runtime_compatibility():
    """给老代码补运行时兼容，同时彻底禁用绘图交互。"""
    import numpy as np
    if not hasattr(np, "int"):
        np.int = int  # type: ignore[attr-defined]

    import matplotlib
    matplotlib.use("Agg", force=True)
    import matplotlib.pyplot as plt

    def _no_show(*args, **kwargs):
        try:
            plt.close("all")
        except Exception:
            pass
        return None

    def _no_pause(*args, **kwargs):
        return None

    def _no_wait(*args, **kwargs):
        return False

    def _no_ion(*args, **kwargs):
        return None

    def _no_ioff(*args, **kwargs):
        return None

    plt.show = _no_show  # type: ignore[assignment]
    plt.pause = _no_pause  # type: ignore[assignment]
    plt.waitforbuttonpress = _no_wait  # type: ignore[assignment]
    plt.ion = _no_ion  # type: ignore[assignment]
    plt.ioff = _no_ioff  # type: ignore[assignment]


def patch_loaded_plot_modules():
    """在 kernel 载入后补丁 util_plot，彻底跳过按键等待。"""
    import matplotlib
    matplotlib.use("Agg", force=True)
    import matplotlib.pyplot as plt

    util_plot = importlib.import_module("util_plot")

    def _show_plot_no_block(hdl):
        try:
            if hasattr(hdl, "hold"):
                hdl.hold = False
            fig = getattr(hdl, "fig", None)
            if fig is not None:
                try:
                    fig.canvas.draw()
                except Exception:
                    pass
                try:
                    fig.canvas.flush_events()
                except Exception:
                    pass
            plt.close("all")
        except Exception:
            pass
        return hdl

    def _close_plot_no_block():
        try:
            plt.close("all")
        except Exception:
            pass
        return None

    util_plot.show_plot = _show_plot_no_block  # type: ignore[attr-defined]
    util_plot.close_plot = _close_plot_no_block  # type: ignore[attr-defined]

    if hasattr(util_plot, "plt"):
        util_plot.plt.show = lambda *a, **k: None  # type: ignore[attr-defined]
        util_plot.plt.pause = lambda *a, **k: None  # type: ignore[attr-defined]
        util_plot.plt.waitforbuttonpress = lambda *a, **k: False  # type: ignore[attr-defined]
        util_plot.plt.close = lambda *a, **k: None  # type: ignore[attr-defined]


def analyze_file_identity(file_stem: str) -> dict:
    nums = NUM_PATTERN.findall(file_stem)
    parsed = []
    for x in nums:
        try:
            parsed.append(float(x))
        except ValueError:
            pass

    temp_value = None
    rep_value = None

    if len(parsed) >= 2:
        last = parsed[-1]
        prev = parsed[-2]
        if float(last).is_integer() and 1 <= int(last) <= 20 and 100 <= prev <= 1200:
            rep_value = int(last)
            temp_value = prev

    if temp_value is None:
        candidates = [x for x in parsed if 100 <= x <= 1200]
        if candidates:
            temp_value = candidates[-1]

    if temp_value is None:
        temp_match = re.search(r"(600|650|700|750|800|850)", file_stem)
        if temp_match:
            temp_value = float(temp_match.group(1))

    if rep_value is None:
        rep_match = re.search(r"-(\d+)$", file_stem)
        if rep_match:
            rep_value = int(rep_match.group(1))

    if temp_value is not None and abs(temp_value - int(temp_value)) < 1e-9:
        label = str(int(temp_value))
    elif temp_value is not None:
        label = str(temp_value)
    else:
        label = file_stem

    return {
        "stem": file_stem,
        "temp_value": temp_value,
        "temp_label": label,
        "replicate": rep_value,
    }


def find_first_data_file(folder: Path) -> Path | None:
    if not folder.exists():
        return None
    files = sorted([p for p in folder.iterdir() if p.is_file() and p.suffix.lower() in {".txt", ".csv"}])
    return files[0] if files else None


def parse_drt_ecm_file(file_path: Path) -> list[tuple[float, float]]:
    data = []
    with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
        for line in f:
            s = line.strip()
            if not s:
                continue
            low = s.lower()
            if (
                "l_self" in low or "l_wire" in low or "r_inf" in low or "r_0" in low
                or ("freq" in low and "gamma" in low)
                or ("tau" in low and "gamma" in low)
            ):
                continue
            parts = re.split(r"[\s,\t;]+", s)
            if len(parts) < 2:
                continue
            try:
                x = float(parts[0])
                y = float(parts[1])
                data.append((x, y))
            except ValueError:
                continue
    return data


def create_summary_workbook(summary_items: list[dict], output_xlsx: Path):
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Alignment, Border, Side, Font

    wb = Workbook()
    ws = wb.active
    ws.title = "汇总"

    if not summary_items:
        wb.save(output_xlsx)
        return

    summary_items.sort(
        key=lambda d: (
            d["temp_value"] is None,
            -(d["temp_value"] if d["temp_value"] is not None else -999999),
            d["label"],
        )
    )

    n = len(summary_items)
    last_col = n * 2
    ws.cell(1, 2, "DRT-Fit-Summary")
    if last_col >= 2:
        ws.merge_cells(start_row=1, start_column=2, end_row=1, end_column=last_col)
    ws["B1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["B1"].font = Font(bold=True, size=12)

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for i, item in enumerate(summary_items):
        col1 = i * 2 + 1
        col2 = col1 + 1
        fill = PatternFill(fill_type="solid", fgColor=HEADER_COLORS[i % len(HEADER_COLORS)])

        ws.cell(2, col1, item["label"])
        ws.cell(2, col2, None)
        ws.cell(3, col1, "freq (Hz)")
        ws.cell(3, col2, "gamma")

        ws.cell(2, col1).fill = fill
        ws.cell(2, col2).fill = fill

        for r, c in [(2, col1), (2, col2), (3, col1), (3, col2)]:
            ws.cell(r, c).border = border
            ws.cell(r, c).alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, (freq, gamma) in enumerate(item["data"], start=4):
            ws.cell(row_idx, col1, freq)
            ws.cell(row_idx, col2, gamma)

    for col in range(1, last_col + 1):
        col_letter = openpyxl.utils.get_column_letter(col)
        ws.column_dimensions[col_letter].width = 13

    wb.save(output_xlsx)


def make_fake_input():
    def fake_input(prompt: str = ""):
        p = str(prompt).lower()
        print(prompt, end="")
        if "area" in p and "cm^2" in p:
            ans = str(AREA_CM2)
        elif "save the results" in p:
            ans = "1"
        elif "input anything" in p:
            ans = "1"
        elif "press enter" in p and "next file" in p:
            ans = "1"
        else:
            ans = "1"
        print(ans)
        return ans
    return fake_input


def snapshot_output_folders(base_dirs: list[Path]) -> dict[str, list[Path]]:
    result = {name: [] for name in KEEP_FOLDERS}
    for base in base_dirs:
        if not base.exists():
            continue
        for name in KEEP_FOLDERS:
            p = base / name
            if p.exists() and p.is_dir():
                result[name].append(p.resolve())
    return result


def find_new_output_folders(before: dict[str, list[Path]], base_dirs: list[Path]) -> dict[str, Path]:
    found: dict[str, Path] = {}
    for base in base_dirs:
        if not base.exists():
            continue
        for name in KEEP_FOLDERS:
            p = base / name
            if p.exists() and p.is_dir() and p.resolve() not in before.get(name, []):
                found[name] = p

    for base in base_dirs:
        if not base.exists():
            continue
        for name in KEEP_FOLDERS:
            if name in found:
                continue
            p = base / name
            if p.exists() and p.is_dir():
                found[name] = p
    return found


def collect_outputs_to_target(found_outputs: dict[str, Path], target_point_dir: Path):
    target_point_dir.mkdir(parents=True, exist_ok=True)
    for folder_name in KEEP_FOLDERS:
        src = found_outputs.get(folder_name)
        dst = target_point_dir / folder_name
        if src and src.exists() and src.is_dir():
            safe_copytree(src, dst)


def ensure_expected_outputs(point_dir: Path):
    missing = [name for name in KEEP_FOLDERS if not (point_dir / name).exists()]
    if missing:
        raise RuntimeError("未生成完整输出文件夹：" + ", ".join(missing))

    drt_ecm = point_dir / "DRT_ECM"
    if not drt_ecm.exists() or find_first_data_file(drt_ecm) is None:
        raise RuntimeError("DRT_ECM 文件夹不存在或里面没有可读取的数据文件。")


def run_eisart_for_one_file(kernel, settings_dir: Path, src_file: Path, final_point_dir: Path) -> dict:
    identity = analyze_file_identity(src_file.stem)

    temp_run_dir = Path(tempfile.mkdtemp(prefix="eisart_run_"))
    copied_input = temp_run_dir / src_file.name
    shutil.copy2(src_file, copied_input)

    old_cwd = Path.cwd()
    old_input = builtins.input

    candidate_output_bases = [
        temp_run_dir,
        settings_dir,
        settings_dir.parent,
        copied_input.parent,
    ]
    before = snapshot_output_folders(candidate_output_bases)

    try:
        os.chdir(settings_dir)
        builtins.input = make_fake_input()
        result = kernel.kernel_main(
            area=AREA_CM2,
            head=str(copied_input),
            save_results=True,
            R_inf_set=None,
            L_set=None,
        )
        if result not in (0, None):
            raise RuntimeError(f"EISART 返回代码异常：{result}")
    finally:
        builtins.input = old_input
        os.chdir(old_cwd)

    found_outputs = find_new_output_folders(before, candidate_output_bases)
    collect_outputs_to_target(found_outputs, final_point_dir)
    ensure_expected_outputs(final_point_dir)

    drt_ecm_file = find_first_data_file(final_point_dir / "DRT_ECM")
    if drt_ecm_file is None:
        raise RuntimeError("未找到 DRT_ECM 输出文件。")

    summary_data = parse_drt_ecm_file(drt_ecm_file)
    if not summary_data:
        raise RuntimeError("DRT_ECM 数据文件已生成，但没有成功解析出 freq/gamma 数据。")

    safe_rmtree(temp_run_dir)

    return {
        "label": identity["temp_label"],
        "temp_value": identity["temp_value"],
        "replicate": identity["replicate"],
        "data": summary_data,
        "source_file": str(src_file),
        "point_dir": str(final_point_dir),
    }


def choose_summary_items(processed_items: list[dict]) -> tuple[list[dict], list[str]]:
    grouped: dict[str, list[dict]] = {}
    for item in processed_items:
        key = item["label"]
        grouped.setdefault(key, []).append(item)

    selected: list[dict] = []
    notes: list[str] = []

    for label, items in grouped.items():
        if len(items) == 1:
            selected.append(items[0])
            continue

        rep1 = [x for x in items if x.get("replicate") == 1]
        if rep1:
            chosen = sorted(rep1, key=lambda d: Path(d["source_file"]).name)[0]
            selected.append(chosen)
            notes.append(f"{label}: 检测到重复数据 {len(items)} 组，仅提取编号 1 -> {Path(chosen['source_file']).name}")
        else:
            chosen = sorted(items, key=lambda d: Path(d["source_file"]).name)[0]
            selected.append(chosen)
            notes.append(f"{label}: 检测到重复数据 {len(items)} 组，但未找到编号 1，已改为提取 {Path(chosen['source_file']).name}")

    return selected, notes


def select_drt_input_files(files: list[Path], manual_mode: bool = False) -> list[Path]:
    if manual_mode:
        return sorted(files, key=lambda p: p.name.lower())

    grouped: dict[str, list[tuple[dict, Path]]] = {}
    for path in files:
        identity = analyze_file_identity(path.stem)
        key = identity["temp_label"]
        grouped.setdefault(key, []).append((identity, path))

    selected: list[Path] = []
    for key, items in grouped.items():
        items.sort(
            key=lambda pair: (
                pair[0]["replicate"] is None,
                pair[0]["replicate"] if pair[0]["replicate"] is not None else 10**9,
                pair[1].name.lower(),
            )
        )
        selected.append(items[0][1])

    selected.sort(
        key=lambda p: (
            analyze_file_identity(p.stem)["temp_value"] is None,
            -(analyze_file_identity(p.stem)["temp_value"] or -999999),
            p.name.lower(),
        )
    )
    return selected


def main():
    try:
        if sys.version_info[:2] != (3, 11):
            raise RuntimeError(f"请使用 Python 3.11 运行本脚本。\n当前版本：{sys.version}")

        print("当前 Python：", sys.executable)
        print("当前版本：", sys.version)

        code_dir = find_eisart_code_dir()
        validate_eisart_code_dir(code_dir)
        patch_eisart_source_compatibility(code_dir)
        monkeypatch_runtime_compatibility()

        selected_files = choose_eis_files()
        if not selected_files:
            print("未选择任何文件。")
            return

        input_files, skipped_inputs = filter_selected_input_files(selected_files)
        if not input_files:
            raise RuntimeError(
                "你选中的文件里没有识别到可处理的原始 EIS 文件。\n\n"
                + "\n".join(skipped_inputs[:20])
            )

        export_base = choose_directory("选择导出文件夹位置")
        if not export_base:
            print("已取消。")
            return

        root_export_dir = export_base / "数据拟合"
        if root_export_dir.exists():
            safe_rmtree(root_export_dir)
        root_export_dir.mkdir(parents=True, exist_ok=True)

        settings_dir = find_settings_dir(code_dir)
        if str(code_dir) not in sys.path:
            sys.path.insert(0, str(code_dir))

        kernel = load_module_from_file("EISART_kernel", code_dir / "EISART_kernel.py")
        kernel.connected_to_gui = False
        patch_loaded_plot_modules()

        processed_items: list[dict] = []
        logs: list[str] = []
        failed: list[str] = []

        total = len(input_files)
        for idx, src_file in enumerate(input_files, start=1):
            print(f"[{idx}/{total}] 正在处理：{src_file.name}")
            point_dir = root_export_dir / src_file.stem
            point_dir.mkdir(parents=True, exist_ok=True)

            try:
                item = run_eisart_for_one_file(
                    kernel=kernel,
                    settings_dir=settings_dir,
                    src_file=src_file,
                    final_point_dir=point_dir,
                )
                processed_items.append(item)
                logs.append(f"成功：{src_file.name}")
                print(f"完成：{src_file.name}")
            except Exception as e:
                err = f"失败：{src_file.name} -> {type(e).__name__}: {e}"
                failed.append(err)
                logs.append(err)
                print(err)
                traceback.print_exc()

        log_file = root_export_dir / "处理日志.txt"
        with open(log_file, "w", encoding="utf-8") as f:
            f.write("EISART 批量处理日志\n")
            f.write(f"面积 area = {AREA_CM2} cm^2\n")
            f.write(f"EISART code dir = {code_dir}\n")
            f.write(f"Python = {sys.executable}\n\n")
            if skipped_inputs:
                f.write("自动跳过的非 EIS 文件：\n")
                for line in skipped_inputs:
                    f.write(line + "\n")
                f.write("\n")
            for line in logs:
                f.write(line + "\n")

        if failed:
            raise RuntimeError(
                "批量处理失败，未继续生成最终汇总表。\n\n" + "\n".join(failed[:10])
            )

        selected_items, summary_notes = choose_summary_items(processed_items)
        summary_xlsx = root_export_dir / "DRT-Fit-Summary.xlsx"
        create_summary_workbook(selected_items, summary_xlsx)

        if summary_notes:
            with open(log_file, "a", encoding="utf-8") as f:
                f.write("\nSummary 选择说明：\n")
                for line in summary_notes:
                    f.write(line + "\n")

        msg = (
            f"批量处理完成。\n\n"
            f"EISART路径：{code_dir}\n"
            f"导出目录：{root_export_dir}\n"
            f"汇总表：{summary_xlsx.name}\n"
            f"日志：{log_file.name}\n"
            f"成功数量：{len(processed_items)}/{len(input_files)}\n"
            f"自动跳过数量：{len(skipped_inputs)}\n"
            f"Summary提取数量：{len(selected_items)}"
        )
        print(msg)
        show_info(msg, "完成")

    except Exception as e:
        traceback.print_exc()
        show_error(f"{type(e).__name__}:\n{e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
