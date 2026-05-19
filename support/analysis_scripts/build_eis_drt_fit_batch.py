from __future__ import annotations

import importlib.util
import os
import re
import shutil
import sys
import tempfile
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


AREA_CM2 = 0.196
REPO_ROOT = Path(__file__).resolve().parents[1]
ROOT = Path(
    os.environ.get(
        "LSV_EIS_DATA_ROOT",
        str(Path.cwd()),
    )
)
OUTPUT_DIR = Path(
    os.environ.get(
        "LSV_EIS_OUTPUT",
        str(Path.cwd() / "processed"),
    )
)
LOCAL_TEMPLATE = REPO_ROOT / "templates" / "文件夹名称-EIS+DRT+拟合.xlsx"
if "EIS_DRT_TEMPLATE" in os.environ:
    TEMPLATE = Path(os.environ["EIS_DRT_TEMPLATE"])
elif LOCAL_TEMPLATE.exists():
    TEMPLATE = LOCAL_TEMPLATE
else:
    raise RuntimeError("EIS+DRT template not found. Set EIS_DRT_TEMPLATE or restore support/templates.")
DRT_RUNNER = REPO_ROOT / "eisart_drt_tool" / "run_eisart_batch_summary_final.py"
LOCAL_DRT_RUNNER = REPO_ROOT / "eisart_drt_tool" / "run_eisart_batch_summary_final.py"
if "DRT_RUNNER" in os.environ:
    DRT_RUNNER = Path(os.environ["DRT_RUNNER"])
elif LOCAL_DRT_RUNNER.exists():
    DRT_RUNNER = LOCAL_DRT_RUNNER
DRT_TOOL_DIR = DRT_RUNNER.parent

TEMP_PALETTES = {
    "600": ["FFFFE699", "FFFFD966", "FFF1C232", "FFBF9000"],
    "650": ["FFE2F0D9", "FFC6E0B4", "FFA9D18E", "FF70AD47"],
    "700": ["FFDDEBF7", "FFBDD7EE", "FF9DC3E6", "FF5B9BD5"],
    "750": ["FFFCE4D6", "FFF8CBAD", "FFF4B183", "FFE69138"],
    "800": ["FFFFCCCC", "FFFF9999", "FFFF6666", "FFC00000"],
    "850": ["FFE4DFEC", "FFD9EAD3", "FFB4A7D6", "FF674EA7"],
}
FALLBACK_COLORS = ["FFE7E6E6", "FFD9D9D9", "FFBFBFBF", "FF808080"]

BLANK_FILL = PatternFill(fill_type=None)
BLANK_BORDER = Border()
BLACK_FONT = Font(color="FF000000", bold=False)
BLACK_BOLD = Font(color="FF000000", bold=True)
CENTER = Alignment(horizontal="center", vertical="center")
THIN = Side(style="thin", color="FF808080")
THIN_BLACK = Side(style="thin", color="FF000000")
THICK = Side(style="medium", color="FF000000")


def load_module(name: str, path: Path):
    spec = importlib.util.spec_from_file_location(name, path)
    if spec is None or spec.loader is None:
        raise RuntimeError(f"Cannot load module: {path}")
    module = importlib.util.module_from_spec(spec)
    sys.modules[name] = module
    spec.loader.exec_module(module)
    return module


def is_excluded(path: Path) -> bool:
    lowered = str(path).lower()
    return "sta" in lowered


def is_impedance_context(path: Path) -> bool:
    lowered = str(path).lower()
    return ("impedance" in lowered or "阻抗" in str(path)) and not is_excluded(path)


def is_eis_source(path: Path) -> bool:
    if path.suffix.lower() != ".txt" or is_excluded(path):
        return False
    return "eis" in path.name.lower() or is_impedance_context(path)


def parse_eis(path: Path) -> list[tuple[float, float, float]]:
    rows: list[tuple[float, float, float]] = []
    for line in path.read_text(encoding="utf-8-sig", errors="ignore").splitlines()[1:]:
        parts = [part.strip() for part in line.split("\t") if part.strip()]
        if len(parts) < 3:
            parts = [part for part in re.split(r"[\s,;]+", line.strip()) if part]
        if len(parts) < 3:
            continue
        try:
            rows.append((float(parts[0]), float(parts[1]), float(parts[2])))
        except ValueError:
            continue
    return rows


def parse_numeric_columns(path: Path, keep_cols: tuple[int, int], negate_second: bool = False) -> list[tuple[float, float]]:
    rows: list[tuple[float, float]] = []
    for line in path.read_text(encoding="utf-8-sig", errors="ignore").splitlines():
        if not line.strip() or line.lstrip().startswith("#"):
            continue
        parts = [part for part in re.split(r"[\s,\t;]+", line.strip()) if part]
        if len(parts) <= max(keep_cols):
            continue
        try:
            first = float(parts[keep_cols[0]])
            second = float(parts[keep_cols[1]])
        except ValueError:
            continue
        rows.append((first, -second if negate_second else second))
    return rows


def extract_temp_and_group(stem: str) -> tuple[str, str]:
    match = re.search(r"^(1?[A-Z])[-_]?(\d{3})", stem)
    if match:
        return match.group(2), match.group(1)
    match = re.search(r"(\d{3})", stem)
    return (match.group(1), "") if match else ("", "")


def safe_sheet_name(path: Path, used: set[str]) -> str:
    base = path.stem[:31]
    name = base
    idx = 1
    while name in used:
        suffix = f"_{idx}"
        name = f"{base[:31 - len(suffix)]}{suffix}"
        idx += 1
    used.add(name)
    return name


def copy_row_style(source: Worksheet, target: Worksheet, source_row: int, target_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        src = source.cell(source_row, col)
        dst = target.cell(target_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)


def prepare_data_sheet(ws: Worksheet, rows: list[tuple[float, float, float]]) -> None:
    needed_last = max(2, len(rows) + 1)
    while ws.max_row < needed_last:
        copy_row_style(ws, ws, ws.max_row, ws.max_row + 1, ws.max_column)
    for row in range(2, max(ws.max_row, needed_last) + 1):
        for col in range(1, 4):
            ws.cell(row, col).value = None
    for idx, (frequency, zre, zim) in enumerate(rows, start=2):
        ws.cell(idx, 1).value = frequency
        ws.cell(idx, 2).value = zre
        ws.cell(idx, 3).value = zim


def has_value(cell) -> bool:
    return cell.value is not None and str(cell.value).strip() != ""


def block_last_row(ws: Worksheet, start: int, end: int, max_row: int) -> int:
    last = 0
    for row in range(1, max_row + 1):
        if any(has_value(ws.cell(row, col)) for col in range(start, end + 1)):
            last = row
    return last


def paint_block(ws: Worksheet, start: int, end: int, last_row: int, color: str, sub_width: int) -> None:
    fill = PatternFill(fill_type="solid", fgColor=color)
    for row in range(1, last_row + 1):
        for col in range(start, end + 1):
            cell = ws.cell(row, col)
            cell.fill = copy(fill)
            cell.font = copy(BLACK_BOLD if row == 1 else BLACK_FONT)
            cell.alignment = CENTER
            left = THIN_BLACK if col == start else THIN
            right = THICK if ((col - start + 1) % sub_width == 0) else THIN
            top = THIN_BLACK if row == 1 else THIN
            bottom = THIN_BLACK if row == last_row else THIN
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)


def clear_columns(ws: Worksheet, max_col: int, max_row: int) -> None:
    for col in range(1, max_col + 1):
        for row in range(1, max_row + 1):
            cell = ws.cell(row, col)
            if isinstance(cell.value, str) and "Zim" in cell.value:
                cell.value = None
            cell.fill = BLANK_FILL
            cell.border = BLANK_BORDER
            cell.font = copy(BLACK_FONT)
            cell.alignment = CENTER


def folder_output_path(folder: Path) -> Path:
    relative = folder.relative_to(ROOT)
    output_name = TEMPLATE.name.replace("文件夹名称", folder.name)
    return OUTPUT_DIR / relative.parent / output_name


def eis_dirs() -> list[Path]:
    folders = []
    for folder in ROOT.rglob("*"):
        if not folder.is_dir() or is_excluded(folder):
            continue
        files = [p for p in folder.glob("*.txt") if is_eis_source(p)]
        if files:
            folders.append(folder)
    return sorted(folders, key=lambda p: str(p))


def setup_drt_runner():
    runner = load_module("run_eisart_batch_summary_final_local", DRT_RUNNER)
    code_dir = runner.find_eisart_code_dir()
    runner.validate_eisart_code_dir(code_dir)
    runner.patch_eisart_source_compatibility(code_dir)
    runner.monkeypatch_runtime_compatibility()
    if str(code_dir) not in sys.path:
        sys.path.insert(0, str(code_dir))
    settings_dir = runner.find_settings_dir(code_dir)
    kernel = runner.load_module_from_file("EISART_kernel", code_dir / "EISART_kernel.py")
    kernel.connected_to_gui = False
    runner.patch_loaded_plot_modules()
    return runner, kernel, settings_dir


def run_drt_for_folder(runner, kernel, settings_dir: Path, folder: Path, temp_export_base: Path) -> Path:
    root_export_dir = temp_export_base / "fit_outputs"
    if root_export_dir.exists():
        shutil.rmtree(root_export_dir, ignore_errors=True)
    root_export_dir.mkdir(parents=True, exist_ok=True)
    files = sorted([p for p in folder.glob("*.txt") if is_eis_source(p)], key=lambda p: p.name)
    for src_file in files:
        point_dir = root_export_dir / src_file.stem
        point_dir.mkdir(parents=True, exist_ok=True)
        runner.run_eisart_for_one_file(kernel, settings_dir, src_file, point_dir)
    return root_export_dir


def record_color(records: list[dict]) -> None:
    group_order: dict[tuple[str, str], int] = {}
    group_seen: dict[tuple[str, str], int] = {}
    for record in records:
        temp, group = extract_temp_and_group(record["path"].stem)
        key = (temp, group)
        if key not in group_order:
            group_order[key] = len([item for item in group_order if item[0] == temp])
        repeat_index = group_seen.get(key, 0)
        group_seen[key] = repeat_index + 1
        palette = TEMP_PALETTES.get(temp, FALLBACK_COLORS)
        record["color"] = palette[(group_order[key] * 3 + repeat_index) % len(palette)]


def build_workbook(folder: Path, drt_root: Path) -> Path:
    source_files = sorted([p for p in folder.glob("*.txt") if is_eis_source(p)], key=lambda p: p.name)
    used_sheet_names: set[str] = set()
    records: list[dict] = []
    for path in source_files:
        stem = path.stem
        records.append(
            {
                "path": path,
                "sheet": safe_sheet_name(path, used_sheet_names),
                "eis": parse_eis(path),
                "ecm": parse_numeric_columns(drt_root / stem / "EIS_ECM" / f"{stem}_eis_ecm.txt", (1, 2), True),
                "drt": parse_numeric_columns(drt_root / stem / "DRT" / f"{stem}_drt.txt", (0, 1), False),
            }
        )
    record_color(records)

    output = folder_output_path(folder)
    output.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(TEMPLATE, output)
    wb = load_workbook(output)

    r_summary = wb.worksheets[0]
    rp_summary = wb.worksheets[1]
    drt_summary = wb.worksheets[2]
    template_ws = wb.worksheets[3]

    for sheet in list(wb.sheetnames):
        if sheet not in {r_summary.title, rp_summary.title, drt_summary.title, template_ws.title}:
            del wb[sheet]

    for record in records:
        ws = wb.copy_worksheet(template_ws)
        ws.title = record["sheet"]
        ws.sheet_properties.tabColor = record["color"]
        prepare_data_sheet(ws, record["eis"])
    del wb[template_ws.title]

    for ws in (r_summary, rp_summary, drt_summary):
        for row in ws.iter_rows():
            for cell in row:
                cell.value = None

    max_eis_points = max((len(record["eis"]) for record in records), default=0)
    max_drt_points = max((len(record["drt"]) for record in records), default=0)
    r_max_row = max_eis_points + 1
    rp_max_row = max_eis_points + 2
    drt_max_row = max_drt_points + 1

    for idx, record in enumerate(records):
        r_col = idx * 4 + 1
        d_col = idx * 2 + 1
        sheet_ref = "'" + record["sheet"].replace("'", "''") + "'"

        r_summary.cell(1, r_col).value = record["sheet"]
        r_summary.cell(1, r_col + 2).value = "EIS fit"
        for row in range(2, r_max_row + 1):
            src_row = row
            r_summary.cell(row, r_col).value = f"={sheet_ref}!B{src_row}*{AREA_CM2}"
            r_summary.cell(row, r_col + 1).value = f"={sheet_ref}!C{src_row}*-{AREA_CM2}"
        for row, (zre_fit, zim_fit) in enumerate(record["ecm"], start=2):
            r_summary.cell(row, r_col + 2).value = zre_fit
            r_summary.cell(row, r_col + 3).value = zim_fit

        rp_summary.cell(1, r_col).value = record["sheet"]
        rp_summary.cell(1, r_col + 2).value = "EIS fit"
        positive_z_rows = [
            (row_index, -zim * AREA_CM2)
            for row_index, (_, _, zim) in enumerate(record["eis"], start=2)
            if -zim * AREA_CM2 > 0
        ]
        reference_row = min(positive_z_rows, key=lambda item: item[1])[0] if positive_z_rows else 2
        fit_positive = [(i + 2, pair[1]) for i, pair in enumerate(record["ecm"]) if pair[1] > 0]
        fit_reference_row = min(fit_positive, key=lambda item: item[1])[0] if fit_positive else 2
        rp_summary.cell(2, r_col).value = f"={get_column_letter(r_col)}{rp_max_row}"
        rp_summary.cell(2, r_col + 2).value = f"={get_column_letter(r_col + 2)}{rp_max_row}"
        for row in range(3, rp_max_row + 1):
            src_row = row - 1
            rp_summary.cell(row, r_col).value = (
                f"='{r_summary.title}'!{get_column_letter(r_col)}{src_row}-"
                f"'{r_summary.title}'!{get_column_letter(r_col)}{reference_row}"
            )
            rp_summary.cell(row, r_col + 1).value = f"='{r_summary.title}'!{get_column_letter(r_col + 1)}{src_row}"
            rp_summary.cell(row, r_col + 2).value = (
                f"='{r_summary.title}'!{get_column_letter(r_col + 2)}{src_row}-"
                f"'{r_summary.title}'!{get_column_letter(r_col + 2)}{fit_reference_row}"
            )
            rp_summary.cell(row, r_col + 3).value = f"='{r_summary.title}'!{get_column_letter(r_col + 3)}{src_row}"

        drt_summary.cell(1, d_col).value = record["sheet"]
        for row, (freq, gamma) in enumerate(record["drt"], start=2):
            drt_summary.cell(row, d_col).value = freq
            drt_summary.cell(row, d_col + 1).value = gamma

    style_summary(wb, records, r_max_row, rp_max_row, drt_max_row)
    wb._sheets = [r_summary, rp_summary, drt_summary] + [wb[record["sheet"]] for record in records]
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.save(output)
    return output


def style_summary(wb, records: list[dict], r_max_row: int, rp_max_row: int, drt_max_row: int) -> None:
    specs = [
        (wb.worksheets[0], 4, 2, r_max_row),
        (wb.worksheets[1], 4, 2, rp_max_row),
        (wb.worksheets[2], 2, 2, drt_max_row),
    ]
    for ws, block_width, sub_width, target_row in specs:
        max_col = len(records) * block_width
        max_row = max(ws.max_row, target_row, 1200)
        clear_columns(ws, max_col, max_row)
        for idx, record in enumerate(records):
            start = idx * block_width + 1
            end = start + block_width - 1
            last = block_last_row(ws, start, end, max_row)
            paint_block(ws, start, end, last, record["color"], sub_width)
        for col in range(1, max_col + 1):
            ws.column_dimensions[get_column_letter(col)].width = 14 if block_width == 4 else 16


def main() -> None:
    folders = eis_dirs()
    print(f"folders\t{len(folders)}")
    runner, kernel, settings_dir = setup_drt_runner()
    built: list[Path] = []
    failures: list[str] = []
    for idx, folder in enumerate(folders, start=1):
        files = sorted([p for p in folder.glob("*.txt") if is_eis_source(p)], key=lambda p: p.name)
        print(f"[{idx}/{len(folders)}]\t{folder.relative_to(ROOT)}\tfiles={len(files)}")
        temp_dir = Path(tempfile.mkdtemp(prefix="eis_drt_batch_"))
        try:
            drt_root = run_drt_for_folder(runner, kernel, settings_dir, folder, temp_dir)
            output = build_workbook(folder, drt_root)
            built.append(output)
            print(f"built\t{output}")
        except Exception as exc:
            failures.append(f"{folder}: {type(exc).__name__}: {exc}")
            print(f"failed\t{folder}\t{type(exc).__name__}: {exc}")
        finally:
            shutil.rmtree(temp_dir, ignore_errors=True)
    print(f"built_total\t{len(built)}")
    if failures:
        print("failures")
        for failure in failures:
            print(failure)
        raise SystemExit(1)


if __name__ == "__main__":
    main()
