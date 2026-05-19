from __future__ import annotations

import re
import shutil
import os
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


AREA_CM2 = 0.196
REPO_ROOT = Path(__file__).resolve().parents[1]
LOCAL_TEMPLATE = REPO_ROOT / "templates" / "文件夹名称-LSV.xlsx"
if "LSV_TEMPLATE" in os.environ:
    TEMPLATE = Path(os.environ["LSV_TEMPLATE"])
elif LOCAL_TEMPLATE.exists():
    TEMPLATE = LOCAL_TEMPLATE
else:
    raise RuntimeError("LSV template not found. Set LSV_TEMPLATE or restore support/templates.")
ROOT = Path(
    os.environ.get(
        "LSV_EIS_DATA_ROOT",
        str(Path.cwd()),
    )
)
OUTPUT_DIR = Path(
    os.environ.get(
        "LSV_EIS_OUTPUT",
        str(Path.cwd() / "processed"),
    )
)

BASE_SHEET = ".txt的名称"
SUMMARY_SHEET = "汇总"

# High-contrast palettes. Same temperature keeps one color family, while
# 1A800 and A800 occupy different contrast bands in the same family.
TEMP_COLORS = {
    "600": ["FFFFD966", "FFFF0000", "FFC00000", "FF7F0000", "FFFFA500", "FFB45F06"],
    "650": ["FFC6E0B4", "FF00B050", "FF548235", "FF375623", "FF92D050", "FF006100"],
    "700": ["FFFFFF00", "FFFFC000", "FFBF9000", "FF7F6000", "FFFFE699", "FFC55A11"],
    "750": ["FFF4B183", "FFE69138", "FFC65911", "FF833C0C", "FFFFC000", "FF7F6000"],
    "800": ["FFFF0000", "FFC00000", "FF800000", "FFE06666", "FFCC4125", "FFA61C00"],
    "850": ["FF9DC3E6", "FF5B9BD5", "FF2F75B5", "FF1F4E79", "FF7030A0", "FF5F497A"],
}
FALLBACK_COLORS = ["FFD9EAD3", "FF6AA84F", "FF274E13", "FFCFE2F3", "FF3D85C6", "FF1C4587"]


def parse_table(path: Path) -> list[tuple[float, float]]:
    lines = path.read_text(encoding="utf-8-sig", errors="ignore").splitlines()
    header = lines[0].lower() if lines else ""
    current_in_ma = "/ma" in header or "i /ma" in header
    out: list[tuple[float, float]] = []
    for line in lines[1:]:
        parts = [part.strip() for part in line.split("\t") if part.strip()]
        if len(parts) < 2:
            continue
        try:
            current = float(parts[1])
            if current_in_ma:
                current /= 1000.0
            out.append((float(parts[0]), current))
        except ValueError:
            continue
    return out


def extract_temp_and_group(stem: str) -> tuple[str, str]:
    match = re.search(r"^(1?[A-Z])(\d{3})", stem)
    if match:
        return match.group(2), match.group(1)
    match = re.search(r"(\d{3})", stem)
    return (match.group(1), "") if match else ("", "")


def safe_sheet_name(path: Path, used: set[str]) -> str:
    base = path.stem[:31]
    name = base
    idx = 1
    while name in used:
        suffix = f"_{idx}"
        name = f"{base[:31 - len(suffix)]}{suffix}"
        idx += 1
    used.add(name)
    return name


def quote_sheet(name: str) -> str:
    return "'" + name.replace("'", "''") + "'"


def font_for_fill(argb: str) -> Font:
    rgb = argb[-6:]
    r, g, b = int(rgb[0:2], 16), int(rgb[2:4], 16), int(rgb[4:6], 16)
    luminance = 0.2126 * r + 0.7152 * g + 0.0722 * b
    return Font(color="FFFFFFFF" if luminance < 120 else "FF000000", bold=True)


def copy_row_style(source: Worksheet, target: Worksheet, source_row: int, target_row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        src = source.cell(source_row, col)
        dst = target.cell(target_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        dst.font = copy(src.font)
        dst.fill = copy(src.fill)
        dst.border = copy(src.border)
        dst.alignment = copy(src.alignment)
        dst.number_format = src.number_format
        dst.protection = copy(src.protection)


def prepare_data_sheet(ws: Worksheet, rows: list[tuple[float, float]]) -> None:
    needed_last = max(2, len(rows) + 1)
    while ws.max_row < needed_last:
        copy_row_style(ws, ws, ws.max_row, ws.max_row + 1, ws.max_column)

    for row in range(2, max(ws.max_row, needed_last) + 1):
        for col in range(1, 9):
            ws.cell(row, col).value = None

    for idx, (potential, current) in enumerate(rows, start=2):
        ws.cell(idx, 1).value = potential
        ws.cell(idx, 2).value = current
        ws.cell(idx, 3).value = f"=B{idx}/-{AREA_CM2}"
        ws.cell(idx, 4).value = f"=A{idx}*1"
        ws.cell(idx, 5).value = f"=C{idx}*D{idx}"
        ws.cell(idx, 7).value = potential
        ws.cell(idx, 8).value = current
    ws["F1"] = "峰值功率密度"
    ws["F2"] = "=MAX(E:E)"


def set_block_fill(ws: Worksheet, start_col: int, end_col: int, color: str, max_row: int) -> None:
    fill = PatternFill(fill_type="solid", fgColor=color)
    font = font_for_fill(color)
    for col in range(start_col, end_col + 1):
        for row in range(1, max_row + 1):
            cell = ws.cell(row, col)
            cell.fill = copy(fill)
            cell.font = copy(font)


def lsv_dirs() -> list[Path]:
    out = []
    for folder in ROOT.rglob("*"):
        if not folder.is_dir() or "性能" not in str(folder):
            continue
        files = [p for p in folder.glob("*.txt") if re.search(r"(?i)(lsv|ipv)", p.name)]
        if files:
            out.append(folder)
    return sorted(out, key=lambda p: str(p))


def output_name(folder: Path) -> str:
    return f"{folder.name}-LSV.xlsx"


def output_path(folder: Path) -> Path:
    # Mirror the exported raw-data directory, but avoid a redundant final folder
    # when the workbook name already carries that folder name.
    relative = folder.relative_to(ROOT)
    return OUTPUT_DIR / relative.parent / output_name(folder)


def build_one(folder: Path) -> tuple[Path, int]:
    output = output_path(folder)
    output.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(TEMPLATE, output)
    wb = load_workbook(output)
    template_ws = wb[BASE_SHEET]

    source_files = sorted(
        [p for p in folder.glob("*.txt") if re.search(r"(?i)(lsv|ipv)", p.name)],
        key=lambda p: p.name,
    )

    files_with_keys = []
    for path in source_files:
        temp, group = extract_temp_and_group(path.stem)
        files_with_keys.append((path, temp, group))

    group_order: dict[tuple[str, str], int] = {}
    for _, temp, group in files_with_keys:
        key = (temp, group)
        if key not in group_order:
            group_order[key] = len([item for item in group_order if item[0] == temp])

    records = []
    used_sheet_names: set[str] = set()
    group_seen: dict[tuple[str, str], int] = {}
    for path, temp, group in files_with_keys:
        repeat_index = group_seen.get((temp, group), 0)
        group_seen[(temp, group)] = repeat_index + 1
        palette = TEMP_COLORS.get(temp, FALLBACK_COLORS)
        color = palette[(group_order[(temp, group)] * 3 + repeat_index) % len(palette)]
        records.append(
            {
                "path": path,
                "sheet": safe_sheet_name(path, used_sheet_names),
                "rows": parse_table(path),
                "color": color,
            }
        )

    for sheet in list(wb.sheetnames):
        if sheet not in {SUMMARY_SHEET, BASE_SHEET}:
            del wb[sheet]

    for record in records:
        ws = wb.copy_worksheet(template_ws)
        ws.title = record["sheet"]
        ws.sheet_properties.tabColor = record["color"]
        prepare_data_sheet(ws, record["rows"])

    del wb[template_ws.title]

    summary = wb[SUMMARY_SHEET]
    for row in summary.iter_rows():
        for cell in row:
            cell.value = None
            cell.fill = PatternFill(fill_type=None)

    max_points = max((len(record["rows"]) for record in records), default=0)
    max_row = max_points + 2
    for idx, record in enumerate(records):
        start_col = idx * 3 + 1
        set_block_fill(summary, start_col, start_col + 2, record["color"], max_row)
        summary.cell(1, start_col).value = record["sheet"]
        summary.cell(2, start_col).value = f"={quote_sheet(record['sheet'])}!F2"
        for row in range(3, max_row + 1):
            src_row = row - 1
            summary.cell(row, start_col).value = f"={quote_sheet(record['sheet'])}!C{src_row}*1"
            summary.cell(row, start_col + 1).value = f"={quote_sheet(record['sheet'])}!D{src_row}*1"
            summary.cell(row, start_col + 2).value = f"={quote_sheet(record['sheet'])}!E{src_row}*1"

    for col in range(1, len(records) * 3 + 1):
        summary.column_dimensions[get_column_letter(col)].width = 14

    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.save(output)
    return output, len(records)


def main() -> None:
    for folder in lsv_dirs():
        output, count = build_one(folder)
        print(f"{folder.name}\t{count}\t{output}")


if __name__ == "__main__":
    main()
